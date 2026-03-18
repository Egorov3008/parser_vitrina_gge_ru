"""
Генерация Excel-отчётов для экспорта проектов
"""

import json
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def get_designer(project: dict) -> str:
    """Извлечь проектировщика из characteristics"""
    chars_raw = project.get('characteristics') or '{}'
    if isinstance(chars_raw, str):
        try:
            chars = json.loads(chars_raw)
        except (json.JSONDecodeError, TypeError):
            chars = {}
    else:
        chars = chars_raw

    return (
        chars.get('egrz:Проектировщик')
        or chars.get('object-designer')
        or 'Не указан'
    )


def _apply_header_style(ws, row_num: int, col_count: int):
    """Применить стиль заголовков"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _auto_width(ws, min_width=10, max_width=50):
    """Автоширина столбцов"""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def generate_full_export(projects: list) -> bytes:
    """Все поля проектов → Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"

    headers = [
        "ID", "№ экспертизы", "Объект", "Регион", "Категория",
        "Проектировщик", "Дата публикации", "URL",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    _apply_header_style(ws, 1, len(headers))

    for row_idx, project in enumerate(projects, 2):
        ws.cell(row=row_idx, column=1, value=project.get('vitrina_id', ''))
        ws.cell(row=row_idx, column=2, value=project.get('expertise_num', ''))
        ws.cell(row=row_idx, column=3, value=project.get('object_name', ''))
        ws.cell(row=row_idx, column=4, value=project.get('region', ''))
        ws.cell(row=row_idx, column=5, value=project.get('category', ''))
        ws.cell(row=row_idx, column=6, value=get_designer(project))
        ws.cell(row=row_idx, column=7, value=project.get('published_at', ''))
        ws.cell(row=row_idx, column=8, value=project.get('url', ''))

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_designers_report(projects: list) -> bytes:
    """Аналитика проектировщиков → Excel (2 листа)"""
    wb = Workbook()

    # Лист 1: Проекты
    ws1 = wb.active
    ws1.title = "Проекты"

    headers1 = ["Регион", "Объект", "Проектировщик", "№ экспертизы", "Дата публикации"]
    for col, header in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=header)
    _apply_header_style(ws1, 1, len(headers1))

    # Собираем данные для аналитики
    designer_stats = {}  # designer -> {count, regions set}

    for row_idx, project in enumerate(projects, 2):
        region = project.get('region', '') or ''
        designer = get_designer(project)

        ws1.cell(row=row_idx, column=1, value=region)
        ws1.cell(row=row_idx, column=2, value=project.get('object_name', ''))
        ws1.cell(row=row_idx, column=3, value=designer)
        ws1.cell(row=row_idx, column=4, value=project.get('expertise_num', ''))
        ws1.cell(row=row_idx, column=5, value=project.get('published_at', ''))

        if designer not in designer_stats:
            designer_stats[designer] = {'count': 0, 'regions': set()}
        designer_stats[designer]['count'] += 1
        if region:
            designer_stats[designer]['regions'].add(region)

    _auto_width(ws1)

    # Лист 2: Аналитика
    ws2 = wb.create_sheet("Аналитика")

    headers2 = ["Проектировщик", "Кол-во проектов", "Регионы"]
    for col, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=header)
    _apply_header_style(ws2, 1, len(headers2))

    # Сортировка по количеству проектов (убывание)
    sorted_designers = sorted(designer_stats.items(), key=lambda x: x[1]['count'], reverse=True)

    for row_idx, (designer, stats) in enumerate(sorted_designers, 2):
        ws2.cell(row=row_idx, column=1, value=designer)
        ws2.cell(row=row_idx, column=2, value=stats['count'])
        ws2.cell(row=row_idx, column=3, value=', '.join(sorted(stats['regions'])))

    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
